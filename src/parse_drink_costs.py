#!/usr/bin/env python3
"""
parse_drink_costs.py — แปลงไฟล์ Excel ต้นทุนเครื่องดื่ม → data/drink-costs.json (seed)

ผลลัพธ์เป็น "ค่าตั้งต้น" ให้ระบบแก้ไขในหน้า dashboard (เพิ่ม/ลบ/แก้เมนู+สูตร).
โครงสร้างคิดแบบ "สูตร = แหล่งความจริง":
  • catalog   : คลังวัตถุดิบกลาง  {ชื่อ: {price(ราคาแพ็ก), qty(ปริมาณ/แพ็ก), unit_cost, unit}}
  • menus     : แต่ละเมนูมี recipe = [{ing, qty}] อ้างอิงคลัง + ราคาขายต่อช่องทาง
  • assumptions: ต้นทุนแฝง 30%, ตัวคูณราคา 1.6, GP/VAT ต่อช่องทาง, ค่าใช้จ่ายคงที่/เดือน

แหล่ง (ค่าเริ่มต้น): import-data/cost/Behind_drinks_cost2025.xlsx
  • ชีต 'สรุปรายการ'  → ราคาขาย 3 ช่องทาง + ต้นทุน/แก้ว (seed)
  • ชีต 'กาแฟ','นม-ชา','คำนวนต้นทุน.' → สูตร + ปริมาณวัตถุดิบ
  • ชีต 'วัตถุดิบ' → ราคากลางวัตถุดิบ (override คลัง)

อัปเดต: แก้ Excel → คัดลอกทับ import-data/cost/ → python src/parse_drink_costs.py → --rebuild
"""

import json
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("ต้องติดตั้ง openpyxl ก่อน:  pip install openpyxl")

PROJECT_ROOT = Path(__file__).parent.parent
DEFAULT_XLSX = PROJECT_ROOT / "import-data" / "cost" / "Behind_drinks_cost2025.xlsx"
OUT_PATH = PROJECT_ROOT / "data" / "drink-costs.json"

OVERHEAD = 0.30            # ต้นทุนแฝง (ค่าน้ำแข็ง/ของเสีย/ฯลฯ)
MARGIN_FACTOR = 1.6       # ตัวคูณแนะนำราคาขาย (กำไร ~60%)
FIXED_COST_MONTHLY = 30000
DAYS_PER_MONTH = 30

CHANNELS = {
    "store":   {"label": "หน้าร้าน", "gp": 0.0,   "vat": 0.0},
    "lineman": {"label": "Lineman",  "gp": 0.30,  "vat": 0.07},
    "shoppee": {"label": "Shoppee",  "gp": 0.32,  "vat": 0.07},
    "grab":    {"label": "Grab",     "gp": 0.251, "vat": 0.07},
}

RECIPE_SHEETS = ["กาแฟ", "นม-ชา", "คำนวนต้นทุน."]
MATERIAL_SHEET = "วัตถุดิบ"
SUMMARY_SHEET = "สรุปรายการ"

COFFEE_KW = ["อเมริกาโน่", "เอสเพรสโซ่", "เอสเปรสโซ่", "ลาเต้", "มอคค่า",
             "คาปู", "กาแฟ", "เอสเย็น", "ช็อตกาแฟ", "ช็อต"]
MATCHA_KW = ["มัทฉะ", "มัทชา"]
PIECE_KW = ["แก้ว", "หลอด", "ฝา", "ถุง", "กระดาษ"]


# ── helpers ──
def num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s in ("", "-", "#VALUE!", "#DIV/0!", "#N/A", "#REF!"):
        return None
    try:
        return float(s.replace(",", ""))
    except ValueError:
        return None


def txt(v):
    return "" if v is None else str(v).strip()


def norm(name):
    return "".join(txt(name).lower().split())


def categorize(name):
    """แยก 5 หมวด: signature (ตั้งเองเท่านั้น) / coffee / tea / milk / soda"""
    n = name.lower()
    if "soda" in n or "โซดา" in name:
        return "soda"
    if any(k in name for k in COFFEE_KW):
        return "coffee"
    if "ชา" in name or "matcha" in n or any(k in name for k in MATCHA_KW):
        return "tea"
    if "นม" in name or "โกโก้" in name or "milk" in n or "cocoa" in n:
        return "milk"
    return "coffee"


def unit_of(name):
    return "ชิ้น" if any(k in name for k in PIECE_KW) else "กรัม/มล."


def r2(x):
    return None if x is None else round(x, 2)


# ── 1) คลังวัตถุดิบ + สูตร (พร้อมปริมาณ) จากชีตสูตร ──
def parse_recipes_and_catalog(wb):
    catalog = {}          # name -> {price, qty, unit_cost, unit}
    recipes = {}          # norm(menu) -> {name, recipe:[{ing,qty}]}

    def add_catalog(name, price, qty):
        if not name or name in catalog:
            return
        if not price or not qty or qty <= 0:
            return
        catalog[name] = {
            "price": round(price, 2), "qty": round(qty, 2),
            "unit_cost": round(price / qty, 5), "unit": unit_of(name),
        }

    for sheet in RECIPE_SHEETS:
        if sheet not in wb.sheetnames:
            continue
        grid = list(wb[sheet].iter_rows(values_only=True))
        nrows = len(grid)
        for r, row in enumerate(grid):
            for c, cell in enumerate(row):
                if txt(cell) != "ลำดับ":
                    continue
                drink = txt(row[c + 8]) if c + 8 < len(row) else ""
                if not drink:
                    continue
                lines = []
                rr = r + 1
                while rr < nrows:
                    cur = grid[rr]
                    label6 = txt(cur[c + 6]) if c + 6 < len(cur) else ""
                    if label6 == "ต้นทุนวัตถุดิบ":
                        break
                    iname = txt(cur[c + 1]) if c + 1 < len(cur) else ""
                    price = num(cur[c + 2]) if c + 2 < len(cur) else None
                    qty = num(cur[c + 3]) if c + 3 < len(cur) else None
                    used = num(cur[c + 6]) if c + 6 < len(cur) else None
                    if iname and used and used > 0:
                        add_catalog(iname, price, qty)
                        lines.append({"ing": iname, "qty": round(used, 2)})
                    rr += 1
                    if rr - r > 30:
                        break
                key = norm(drink)
                if lines and key not in recipes:
                    recipes[key] = {"name": drink, "recipe": lines}

    # ชีตวัตถุดิบ (master) — override ราคากลางที่สะอาดกว่า
    if MATERIAL_SHEET in wb.sheetnames:
        for row in wb[MATERIAL_SHEET].iter_rows(values_only=True):
            if not row or txt(row[0]) in ("", "ลำดับ", "วัตถุดิบ"):
                continue
            name = txt(row[1]) if len(row) > 1 else ""
            price = num(row[2]) if len(row) > 2 else None
            qty = num(row[3]) if len(row) > 3 else None
            if name and price and qty and qty > 0:
                catalog[name] = {
                    "price": round(price, 2), "qty": round(qty, 2),
                    "unit_cost": round(price / qty, 5), "unit": unit_of(name),
                }

    return recipes, catalog


def match_recipe(menu_name, recipes):
    return recipes.get(norm(menu_name))


# ── 2) ราคาขาย + ต้นทุน seed จากชีต 'สรุปรายการ' ──
def parse_summary(wb, recipes):
    grid = list(wb[SUMMARY_SHEET].iter_rows(values_only=True))
    menus, seen = [], set()
    idx = 0
    for row in grid:
        name = txt(row[27]) if len(row) > 27 else ""
        if not name or name == "ชื่อเมนู":
            continue
        price_store = num(row[28]) if len(row) > 28 else None
        price_lineman = num(row[15]) if len(row) > 15 else None
        price_shoppee = num(row[2]) if len(row) > 2 else None
        seed_cost = num(row[29]) if len(row) > 29 else None
        if price_store is None and price_lineman is None and price_shoppee is None:
            continue
        if name in seen:
            continue
        seen.add(name)
        idx += 1
        rec = match_recipe(name, recipes)
        menus.append({
            "id": "m" + str(idx),
            "name": name,
            "category": categorize(name),
            "recipe": rec["recipe"] if rec else [],
            "prices": {"store": r2(price_store), "lineman": r2(price_lineman),
                       "shoppee": r2(price_shoppee), "grab": None},
            "seed_cost_cup": r2(seed_cost),
        })
    return menus


# ── main ──
def parse(xlsx_path: Path) -> dict:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    recipes, catalog = parse_recipes_and_catalog(wb)
    menus = parse_summary(wb, recipes)
    return {
        "source": xlsx_path.name,
        "parsed_at": datetime.now().isoformat(timespec="seconds"),
        "assumptions": {
            "overhead": OVERHEAD,
            "margin_factor": MARGIN_FACTOR,
            "fixed_cost_monthly": FIXED_COST_MONTHLY,
            "days_per_month": DAYS_PER_MONTH,
            "channels": CHANNELS,
        },
        "catalog": catalog,
        "menus": menus,
    }


def main():
    sys.stdout.reconfigure(encoding="utf-8")
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not xlsx.exists():
        sys.exit(f"ไม่พบไฟล์: {xlsx}")
    data = parse(xlsx)
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    n_rec = sum(1 for m in data["menus"] if m["recipe"])
    print(f"[drink-costs] เขียน {OUT_PATH.name}: {len(data['menus'])} เมนู "
          f"({n_rec} มีสูตร) · คลังวัตถุดิบ {len(data['catalog'])} รายการ")


if __name__ == "__main__":
    main()
